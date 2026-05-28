"""Build Concrecia commercial proposal template (.xlsx)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image as XLImage

# ---------- BRAND TOKENS ----------
NAVY = "0A1F5C"
NAVY_DEEP = "050F2E"
YELLOW = "F5C518"
YELLOW_DARK = "D4A912"
YELLOW_SOFT = "FFF7D1"
PAPER = "FAFAF8"
STONE_100 = "EAE8E2"
STONE_200 = "D5D2C9"
STONE_300 = "B5B1A6"
STONE_500 = "5E5B53"
STONE_700 = "232220"
STONE_900 = "0E0E0D"
WHITE = "FFFFFF"
DIM_TEXT = "8C887D"

DISP = "Arial Black"
BODY = "Arial"
MONO = "Consolas"

def F(size=10, bold=False, color=STONE_900, name=BODY, italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

# Borders
thin = Side(style="thin", color=STONE_200)
thin_dark = Side(style="thin", color=STONE_300)
thick_navy = Side(style="medium", color=NAVY)
thick_yellow = Side(style="medium", color=YELLOW)

# Fills
navy_fill = PatternFill("solid", fgColor=NAVY)
navy_deep_fill = PatternFill("solid", fgColor=NAVY_DEEP)
yellow_fill = PatternFill("solid", fgColor=YELLOW)
yellow_soft_fill = PatternFill("solid", fgColor=YELLOW_SOFT)
paper_fill = PatternFill("solid", fgColor=PAPER)
stone_fill = PatternFill("solid", fgColor=STONE_100)
white_fill = PatternFill("solid", fgColor=WHITE)

# Alignments
left = Alignment(horizontal="left", vertical="center", indent=1)
right = Alignment(horizontal="right", vertical="center", indent=1)
center = Alignment(horizontal="center", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

# ---------- WORKBOOK ----------
wb = Workbook()
ws = wb.active
ws.title = "Proposta"

ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 100

# Page setup — A4 portrait, fit em UMA página
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True   # crítico: sem isso o Excel ignora fit-to-page
ws.page_setup.scale = None  # quando fitToPage=True, scale fica auto
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = False
ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3,
                              header=0.0, footer=0.0)

# Column widths — total ~90 chars ≈ 6.6" (cabe folgado em A4 portrait com margem)
widths = {'A': 28, 'B': 10, 'C': 8, 'D': 11, 'E': 14, 'F': 18}
for c, w in widths.items():
    ws.column_dimensions[c].width = w

def merge(rng, value=None, font=None, fill=None, align=None, border=None):
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    cell = ws[top_left]
    if value is not None: cell.value = value
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if align is not None: cell.alignment = align
    if border is not None: cell.border = border

def style(addr, value=None, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws[addr]
    if value is not None: cell.value = value
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if align is not None: cell.alignment = align
    if border is not None: cell.border = border
    if num_fmt is not None: cell.number_format = num_fmt
    return cell

def fill_range(rng, fill):
    for row in ws[rng]:
        for c in row:
            c.fill = fill

def border_row(row_num, **kwargs):
    b = Border(**kwargs)
    for col in "ABCDEF":
        ws[f"{col}{row_num}"].border = b

# ============================================================
# ROW 1 — Navy bar
# ============================================================
ws.row_dimensions[1].height = 6
merge("A1:F1", fill=navy_fill)

# Thin yellow accent under navy bar
ws.row_dimensions[2].height = 3
merge("A2:B2", fill=yellow_fill)

# ============================================================
# ROW 3 — Logo oficial (PNG embutido) + subtitle inline
# ============================================================
ws.row_dimensions[3].height = 65
merge("A3:F3")  # área de respiro do logo

# Embute a logo oficial — 80x80 px (cabe folgado em row 3 = 87px)
logo_path = r"C:\Users\Rafael\Meu Drive\concrecia_drive\marca\logo-concrecia.png"
logo_img = XLImage(logo_path)
logo_img.width = 80
logo_img.height = 80
ws.add_image(logo_img, "A3")

# ROW 4 — Subtitle institucional
ws.row_dimensions[4].height = 14
merge("A4:F4",
      value="CONCRETO USINADO   ·   RIBEIRÃO PRETO / SP",
      font=F(name=MONO, size=8, color=STONE_500),
      align=Alignment(horizontal="right", vertical="center", indent=1))
border_row(4, bottom=thick_navy)

# ROW 5 — spacer
ws.row_dimensions[5].height = 12

# ============================================================
# ROWS 6-7 — META: Proposta Nº / Emissão / Validade
# ============================================================
labels = [("A6:B6", "PROPOSTA Nº"), ("C6:D6", "EMISSÃO"), ("E6:F6", "VALIDADE")]
for rng, txt in labels:
    merge(rng, value=txt,
          font=F(size=7, bold=True, color=DIM_TEXT, name=MONO),
          align=Alignment(horizontal="left", vertical="bottom", indent=1))
ws.row_dimensions[6].height = 14

values = [("A7:B7", "2026-0142"), ("C7:D7", "25/Mai/2026"), ("E7:F7", "30/Mai/2026")]
for rng, txt in values:
    merge(rng, value=txt,
          font=F(size=13, bold=True, color=NAVY, name=MONO),
          align=Alignment(horizontal="left", vertical="center", indent=1))
ws.row_dimensions[7].height = 22

# bottom rule
for col in "ABCDEF":
    ws[f"{col}7"].border = Border(bottom=thin_dark)

# ROW 8 — spacer
ws.row_dimensions[8].height = 14

# ============================================================
# ROW 9-10 — Title block
# ============================================================
merge("A9:F9",
      value="PROPOSTA COMERCIAL  ·  FORNECIMENTO DE CONCRETO USINADO",
      font=F(size=8, bold=True, color=YELLOW_DARK, name=MONO),
      align=Alignment(horizontal="left", vertical="center", indent=1))
ws.row_dimensions[9].height = 14

merge("A10:F10",
      value="Ritz Golf — Av. Luiz Eduardo de Toledo Prado",
      font=F(name=DISP, size=18, bold=True, color=NAVY),
      align=Alignment(horizontal="left", vertical="center", indent=1))
ws.row_dimensions[10].height = 26

# ROW 11 — spacer
ws.row_dimensions[11].height = 10

# ============================================================
# ROW 12-13 — Cliente / Obra / Escopo
# ============================================================
meta_labels = [("A12:B12", "CLIENTE"), ("C12:D12", "OBRA"), ("E12:F12", "ESCOPO")]
for rng, txt in meta_labels:
    merge(rng, value=txt,
          font=F(size=7, bold=True, color=DIM_TEXT, name=MONO),
          align=Alignment(horizontal="left", vertical="bottom", indent=1))
ws.row_dimensions[12].height = 14
border_row(12, top=thin)

meta_vals = [
    ("A13:B13", "3BS Construtora e Incorporadora Ltda"),
    ("C13:D13", "Av. Luiz Eduardo de Toledo Prado · Ribeirão Preto / SP"),
    ("E13:F13", "5 traços + bombeamento · NBR 12.655"),
]
for rng, txt in meta_vals:
    merge(rng, value=txt,
          font=F(size=10, bold=True, color=STONE_900),
          align=left_wrap)
ws.row_dimensions[13].height = 24
border_row(13, bottom=thin)

# ROW 14 — spacer
ws.row_dimensions[14].height = 14

# ============================================================
# ROW 15 — Section 01 label
# ============================================================
style("A15", value="01 / 03",
      font=F(size=8, bold=True, color=YELLOW_DARK, name=MONO),
      align=left)
merge("B15:F15", value="FORNECIMENTO DE CONCRETO",
      font=F(name=DISP, size=12, bold=True, color=NAVY),
      align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[15].height = 20

# ============================================================
# ROW 16 — Table header
# ============================================================
headers = [
    ("A16", "ESPECIFICAÇÃO", left),
    ("B16", "SLUMP (cm)", center),
    ("C16", "BRITA", center),
    ("D16", "VOLUME (m³)", right),
    ("E16", "VALOR / m³", right),
    ("F16", "SUBTOTAL", right),
]
for addr, txt, al in headers:
    style(addr, value=txt,
          font=F(size=8, bold=True, color=WHITE, name=MONO),
          fill=navy_fill, align=al,
          border=Border(bottom=thick_yellow))
ws.row_dimensions[16].height = 24

# ============================================================
# ROWS 17-21 — Concrete rows (editable example data)
# ============================================================
concrete = [
    ("Fck 25,0 MPa · NBR 12.1.6", "12 ± 2", 1, 881.0, 445.92),
    ("Fck 35,0 MPa · NBR 12.1.6", "16 ± 3", 1, 794.0, 519.51),
    ("Fck 35,0 MPa · NBR 12.1.6", "22 ± 3", 0, 680.0, 539.66),
    ("Fck 40,0 MPa · NBR 12.1.6", "12 ± 2", 1, 1823.0, 536.54),
    ("Fck 40,0 MPa · NBR 12.1.6", "14 ± 2", 1, 1255.0, 547.29),
]
for i, (spec, slump, brita, vol, val) in enumerate(concrete):
    r = 17 + i
    ws.row_dimensions[r].height = 18
    style(f"A{r}", value=spec,
          font=F(size=10, bold=True, color=NAVY, name=MONO),
          align=left,
          border=Border(bottom=thin))
    style(f"B{r}", value=slump,
          font=F(size=10, color=STONE_900, name=MONO),
          align=center, border=Border(bottom=thin))
    style(f"C{r}", value=brita,
          font=F(size=10, color=STONE_900, name=MONO),
          align=center, border=Border(bottom=thin))
    style(f"D{r}", value=vol, num_fmt='#,##0.0',
          font=F(size=10, color=STONE_900, name=MONO),
          align=right, border=Border(bottom=thin))
    style(f"E{r}", value=val, num_fmt='"R$" #,##0.00',
          font=F(size=10, bold=True, color=NAVY, name=MONO),
          align=right, border=Border(bottom=thin))
    style(f"F{r}", value=f"=D{r}*E{r}", num_fmt='"R$" #,##0.00',
          font=F(size=10, bold=True, color=NAVY, name=MONO),
          fill=paper_fill, align=right, border=Border(bottom=thin))

# ============================================================
# ROW 22 — Bombeamento row (volume = soma das concretagens)
# ============================================================
r = 22
ws.row_dimensions[r].height = 18
style(f"A{r}", value="Serviço de Bombeamento",
      font=F(size=10, bold=True, color=STONE_900, name=MONO),
      fill=stone_fill, align=left,
      border=Border(top=thin_dark, bottom=thin))
for col in "BC":
    style(f"{col}{r}", value="—",
          font=F(size=10, color=DIM_TEXT, name=MONO),
          fill=stone_fill, align=center,
          border=Border(top=thin_dark, bottom=thin))
style(f"D{r}", value=f"=SUM(D17:D21)", num_fmt='#,##0.0',
      font=F(size=10, color=STONE_900, name=MONO),
      fill=stone_fill, align=right,
      border=Border(top=thin_dark, bottom=thin))
style(f"E{r}", value=75.00, num_fmt='"R$" #,##0.00',
      font=F(size=10, bold=True, color=NAVY, name=MONO),
      fill=stone_fill, align=right,
      border=Border(top=thin_dark, bottom=thin))
style(f"F{r}", value=f"=D{r}*E{r}", num_fmt='"R$" #,##0.00',
      font=F(size=10, bold=True, color=NAVY, name=MONO),
      fill=stone_fill, align=right,
      border=Border(top=thin_dark, bottom=thin))

# ============================================================
# ROW 23 — TOTAL row (dark)
# ============================================================
r = 23
ws.row_dimensions[r].height = 30
merge(f"A{r}:B{r}", value="VALOR TOTAL DA PROPOSTA",
      font=F(size=9, bold=True, color=YELLOW, name=MONO),
      fill=navy_fill, align=Alignment(horizontal="left", vertical="center", indent=1))
merge(f"C{r}:D{r}", value=f"=SUM(D17:D21)",
      font=F(size=10, bold=True, color=YELLOW, name=MONO),
      fill=navy_fill, align=Alignment(horizontal="right", vertical="center"))
ws[f"C{r}"].number_format = '#,##0.0" m³"'
merge(f"E{r}:F{r}", value=f"=SUM(F17:F22)",
      font=F(name=DISP, size=15, bold=True, color=WHITE),
      fill=navy_fill,
      align=Alignment(horizontal="right", vertical="center", indent=0))
ws[f"E{r}"].number_format = '"R$" #,##0.00'
# fill the merged area
for col in "ABCDEF":
    ws[f"{col}{r}"].fill = navy_fill

# ROW 24 — spacer
ws.row_dimensions[24].height = 14

# ============================================================
# ROW 25 — Section 02
# ============================================================
style("A25", value="02 / 03",
      font=F(size=8, bold=True, color=YELLOW_DARK, name=MONO),
      align=left)
merge("B25:F25", value="CONDIÇÕES OPERACIONAIS",
      font=F(name=DISP, size=12, bold=True, color=NAVY),
      align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[25].height = 20

# ============================================================
# ROW 26 — Sub-headers (Bombeamento | Taxas)
# ============================================================
merge("A26:C26", value="SERVIÇO DE BOMBEAMENTO",
      font=F(size=8, bold=True, color=NAVY, name=MONO),
      fill=paper_fill, align=left)
merge("D26:F26", value="TAXAS EM HORÁRIOS EXTRAORDINÁRIOS",
      font=F(size=8, bold=True, color=NAVY, name=MONO),
      fill=paper_fill, align=left)
ws.row_dimensions[26].height = 18

# Yellow accent borders on left side of cards
ws["A26"].border = Border(left=thick_yellow, top=thin)
ws["D26"].border = Border(left=thick_yellow, top=thin)
for col in "BCF":
    pass
ws["B26"].border = Border(top=thin)
ws["C26"].border = Border(top=thin)
ws["E26"].border = Border(top=thin)
ws["F26"].border = Border(top=thin)

bombeamento = [
    ("Bomba Lança · mínimo 20 m³", "R$ 1.500,00"),
    ("Excedente Bomba Lança", "R$ 75,00 / m³"),
    ("Bomba Mangote · mínimo 20 m³", "R$ 1.300,00"),
    ("Excedente Bomba Mangote", "R$ 65,00 / m³"),
]
taxas = [
    ("Segunda a sexta · após 16h", "+ 10,0 %"),
    ("Sábado", "+ 30,0 %"),
    ("Domingo e feriado", "+ 50,0 %"),
    ("Volume mínimo por entrega", "5,0 m³"),
]

for i in range(4):
    r = 27 + i
    ws.row_dimensions[r].height = 15
    # Bombeamento
    merge(f"A{r}:B{r}", value=bombeamento[i][0],
          font=F(size=9, color=STONE_500),
          fill=paper_fill, align=left)
    style(f"C{r}", value=bombeamento[i][1],
          font=F(size=9, bold=True, color=NAVY, name=MONO),
          fill=paper_fill, align=right)
    ws[f"A{r}"].border = Border(left=thick_yellow,
                                bottom=Side(style="dotted", color=STONE_200))
    for col in "BC":
        ws[f"{col}{r}"].border = Border(bottom=Side(style="dotted", color=STONE_200))
    # Taxas
    merge(f"D{r}:E{r}", value=taxas[i][0],
          font=F(size=9, color=STONE_500),
          fill=paper_fill, align=left)
    style(f"F{r}", value=taxas[i][1],
          font=F(size=9, bold=True, color=NAVY, name=MONO),
          fill=paper_fill, align=right)
    ws[f"D{r}"].border = Border(left=thick_yellow,
                                bottom=Side(style="dotted", color=STONE_200))
    for col in "EF":
        ws[f"{col}{r}"].border = Border(bottom=Side(style="dotted", color=STONE_200))

# Footer note (row 31)
r = 31
ws.row_dimensions[r].height = 16
merge(f"A{r}:C{r}", value="Volume mínimo de 20 m³ por período de 4 horas.",
      font=F(size=8, color=STONE_500, italic=True),
      fill=paper_fill, align=left)
merge(f"D{r}:F{r}", value="Acréscimo sobre concreto e bombeamento.",
      font=F(size=8, color=STONE_500, italic=True),
      fill=paper_fill, align=left)
ws[f"A{r}"].border = Border(left=thick_yellow, bottom=thin)
ws[f"D{r}"].border = Border(left=thick_yellow, bottom=thin)
for col in "BCEF":
    ws[f"{col}{r}"].border = Border(bottom=thin)

# ROW 32 — spacer
ws.row_dimensions[32].height = 14

# ============================================================
# ROW 33 — Section 03
# ============================================================
style("A33", value="03 / 03",
      font=F(size=8, bold=True, color=YELLOW_DARK, name=MONO),
      align=left)
merge("B33:F33", value="DISPOSIÇÕES GERAIS",
      font=F(name=DISP, size=12, bold=True, color=NAVY),
      align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[33].height = 20

# ============================================================
# ROWS 34-38 — 5 conditions on navy_deep background
# ============================================================
conditions = [
    ("01", "PAGAMENTO", "28 D.D. — cadastro sujeito a análise de crédito."),
    ("02", "VALIDADE", "30/05/2026. Após esta data, valores sujeitos a reajuste."),
    ("03", "NORMA", "Dosagem conforme NBR 12.655, com controle tecnológico por lote."),
    ("04", "PROGRAMAÇÃO", "Antecedência mínima de 24 h úteis para garantia de janela operacional."),
    ("05", "ACEITE", "A contratação do serviço implica aceitação integral destas condições."),
]
for i, (num, lbl, text) in enumerate(conditions):
    r = 34 + i
    ws.row_dimensions[r].height = 15
    style(f"A{r}", value=num,
          font=F(size=8, bold=True, color=YELLOW, name=MONO),
          fill=navy_deep_fill,
          align=center)
    style(f"B{r}", value=lbl,
          font=F(size=8, bold=True, color=WHITE, name=MONO),
          fill=navy_deep_fill, align=left)
    merge(f"C{r}:F{r}", value=text,
          font=F(size=9, color="D5D2C9"),
          fill=navy_deep_fill, align=left_wrap)
    for col in "CDEF":
        ws[f"{col}{r}"].fill = navy_deep_fill

# ROW 39 — spacer
ws.row_dimensions[39].height = 14

# ============================================================
# ROW 40-42 — Contatos
# ============================================================
merge("A40:F40", value="CONTATOS",
      font=F(size=7, bold=True, color=DIM_TEXT, name=MONO),
      align=Alignment(horizontal="left", vertical="bottom", indent=1))
ws.row_dimensions[40].height = 14
border_row(40, top=thin)

merge("A41:C41", value="Romulo Ferrari",
      font=F(name=DISP, size=12, bold=True, color=NAVY),
      align=left)
merge("D41:F41", value="Junior",
      font=F(name=DISP, size=12, bold=True, color=NAVY),
      align=left)
ws.row_dimensions[41].height = 18

merge("A42:C42", value="+55 (16) 99220-1151  ·  romulo@jjconcreto.com.br",
      font=F(size=9, color=STONE_500),
      align=left)
merge("D42:F42", value="junior@concrecia.com.br",
      font=F(size=9, color=STONE_500),
      align=left)
ws.row_dimensions[42].height = 16

# ============================================================
# ROWS 43-46 — Footer institucional em 3 linhas
# ============================================================
# spacer
ws.row_dimensions[43].height = 8

# Linha 1 — Razão social + CNPJ (âncora institucional)
merge("A44:F44",
      value="CONCRECIA CONCRETO LTDA   ·   CNPJ 11.574.376/0001-26",
      font=F(size=8, bold=True, color=NAVY, name=BODY),
      align=Alignment(horizontal="center", vertical="center"))
ws.row_dimensions[44].height = 16
border_row(44, top=Side(style="medium", color=NAVY))

# Linha 2 — Endereço (info secundária)
merge("A45:F45",
      value="Rod. Antonio Machado Sant'anna, km 7   ·   Ribeirão Preto / SP",
      font=F(size=8, color=STONE_500, name=BODY),
      align=Alignment(horizontal="center", vertical="center"))
ws.row_dimensions[45].height = 14

# Linha 3 — Tagline (assinatura da marca)
merge("A46:F46",
      value="Construindo confiança desde 2009",
      font=F(size=8, italic=True, color=YELLOW_DARK, name=BODY),
      align=Alignment(horizontal="center", vertical="center"))
ws.row_dimensions[46].height = 16

# Print area
ws.print_area = "A1:F46"

# Save
import os
out = r"C:\Users\Rafael\Downloads\Template-Proposta-Comercial-Concrecia.xlsx"
# limpa versão antiga v2 se existir
v2 = r"C:\Users\Rafael\Downloads\Template-Proposta-Comercial-Concrecia-v2.xlsx"
if os.path.exists(v2):
    try: os.remove(v2)
    except: pass
wb.save(out)
print(f"Saved: {out}")
